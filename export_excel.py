import json, urllib.request, urllib.error, base64, time
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LOCATIONS_MAP = {1514: "Sucursal Bellas Artes", 2547: "Sucursal Providencia", 7705: "Sucursal Ñuñoa"}
API_BASE = "https://agendapro.com/api/views/admin/v2/calendar"
RANGE_START = date(2026, 1, 30)
RANGE_END = date(2026, 3, 31)


def daily_chunks(start, end):
    """Yield (day, day) tuples for each date in [start, end]."""
    cursor = start
    while cursor <= end:
        yield cursor, cursor
        cursor += timedelta(days=1)

# --- Auth ---
with open("auth-state.json") as f:
    auth_state = json.load(f)
token = next(c["value"] for c in auth_state["cookies"] if c["name"] == "ap_cognito_authorization")
token = token.replace("Bearer ", "", 1)

def api_get(path):
    req = urllib.request.Request(
        f"{API_BASE}/{path}",
        headers={
            "Accept": "application/json",
            "Authorization": token,
            "Origin": "https://app.agendapro.com",
            "Referer": "https://app.agendapro.com/",
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Expose-Headers": "Authorization",
            "From": base64.b64encode(b"https://app.agendapro.com/bookings").decode(),
        },
    )
    with urllib.request.urlopen(req) as res:
        return json.loads(res.read())


# --- Token expiry pre-check ---
def check_token_expiry(jwt_token):
    """Decode JWT payload and warn/abort if token is near expiry."""
    try:
        payload_b64 = jwt_token.split(".")[1]
        # Fix padding
        payload_b64 += "=" * (4 - len(payload_b64) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        exp = payload.get("exp")
        if exp:
            remaining = exp - time.time()
            if remaining < 300:
                print(f"ERROR: Token expires in {int(remaining)}s (< 5 min). Aborting.")
                raise SystemExit(1)
            elif remaining < 3600:
                print(f"WARNING: Token expires in {int(remaining // 60)} minutes.")
            else:
                print(f"  Token valid for {int(remaining // 3600)}h {int((remaining % 3600) // 60)}m")
    except (IndexError, json.JSONDecodeError, ValueError):
        print("WARNING: Could not decode token expiry. Proceeding anyway.")

check_token_expiry(token)


# --- Retry wrapper ---
def api_get_with_retry(path, max_retries=3):
    """Call api_get with retry on 429 / 5xx errors, exponential backoff."""
    for attempt in range(max_retries + 1):
        try:
            return api_get(path)
        except urllib.error.HTTPError as e:
            if e.code == 429 or e.code >= 500:
                if attempt < max_retries:
                    wait = 2 ** (attempt + 1)
                    print(f"    HTTP {e.code}, retrying in {wait}s (attempt {attempt + 1}/{max_retries})...")
                    time.sleep(wait)
                else:
                    raise
            else:
                raise


# --- Pagination support ---
def fetch_all_bookings(location_id, start, end):
    """Fetch all pages of bookings for a location/date range, merging calendar_users_events."""
    path = f"bookings?start={start}&end={end}&location_id={location_id}&time_resource=false&per_page=100&page=1"
    data = api_get_with_retry(path)
    all_users_events = data.get("calendar_users_events", [])
    total_pages = data.get("total_pages", 1)

    for page in range(2, total_pages + 1):
        time.sleep(0.3)
        page_path = f"bookings?start={start}&end={end}&location_id={location_id}&time_resource=false&per_page=100&page={page}"
        page_data = api_get_with_retry(page_path)
        all_users_events.extend(page_data.get("calendar_users_events", []))

    data["calendar_users_events"] = all_users_events
    return data


# --- Fetch data ---
print("Fetching locations...")
locations = api_get_with_retry("locations?per_page=8&search_key=&page=1")["locations"]
print(f"  Found {len(locations)} locations")

chunks = list(daily_chunks(RANGE_START, RANGE_END))
total_requests = len(locations) * len(chunks)
request_num = 0

all_reserved = {}
all_blocked = {}

for loc in locations:
    lid = loc["value"]
    name = loc["label"]

    reserved_rows = []
    blocked_rows = []
    seen_reserved_ids = set()
    seen_blocked_ids = set()

    for chunk_start, chunk_end in chunks:
        request_num += 1
        print(f"  [{request_num}/{total_requests}] {name}: {chunk_start} to {chunk_end}")
        data = fetch_all_bookings(lid, chunk_start, chunk_end)

        for user in data["calendar_users_events"]:
            prof_name = f"{user['first_name']} {user['last_name']}"
            for ev in user["events"]:
                if ev["type"] in ("RESERVED", "CONFIRMED", "ATTENDED", "WAITLISTED"):
                    b = ev.get("booking", {})
                    ev_id = b.get("id", ev["id"])
                    if ev_id in seen_reserved_ids:
                        continue
                    seen_reserved_ids.add(ev_id)
                    client = b.get("client", {})
                    service = b.get("service", {})
                    reserved_rows.append({
                        "Booking ID": ev_id,
                        "Profesional": prof_name,
                        "Servicio": service.get("name", ev.get("title", "")),
                        "Inicio": ev["start"],
                        "Fin": ev["end"],
                        "Duracion (min)": service.get("duration", ""),
                        "Cliente": f"{client.get('first_name', '')} {client.get('last_name', '')}".strip(),
                        "Email": client.get("email", ""),
                        "Telefono": client.get("phone", ""),
                        "Precio": b.get("price", ""),
                        "Monto": b.get("amount", ""),
                        "Estado Pago": b.get("payment_status", ""),
                        "Cliente Nuevo": "Si" if client.get("is_new_client") else "No",
                        "Tags": ", ".join(b.get("tags", [])),
                        "Comentario": b.get("comment", ""),
                        "Estado": ev["type"],
                    })
                else:
                    ev_id = ev["id"]
                    if ev_id in seen_blocked_ids:
                        continue
                    seen_blocked_ids.add(ev_id)
                    blocked_rows.append({
                        "Event ID": ev_id,
                        "Profesional": prof_name,
                        "Tipo": ev["type"],
                        "Titulo": ev.get("title", ""),
                        "Descripcion": ev.get("description", ""),
                        "Inicio": ev["start"],
                        "Fin": ev["end"],
                    })

        time.sleep(0.3)

    # Sort rows by start date
    reserved_rows.sort(key=lambda r: r["Inicio"])
    blocked_rows.sort(key=lambda r: r["Inicio"])

    all_reserved[lid] = reserved_rows
    all_blocked[lid] = blocked_rows

# --- Styles ---
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
alt_fill = PatternFill("solid", fgColor="D6E4F0")

def write_sheet(ws, headers, rows, location_name):
    ws.title = location_name[:31]  # Excel sheet name limit
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(h, ""))
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, h in enumerate(headers, 1):
        max_len = len(h)
        for row_idx in range(2, len(rows) + 2):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_len = max(max_len, min(len(val), 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

# --- Write Reserved Excel ---
reserved_headers = [
    "Booking ID", "Profesional", "Servicio", "Inicio", "Fin", "Duracion (min)",
    "Cliente", "Email", "Telefono", "Precio", "Monto", "Estado Pago",
    "Cliente Nuevo", "Tags", "Comentario", "Estado",
]

wb_res = Workbook()
wb_res.remove(wb_res.active)
for loc in locations:
    lid = loc["value"]
    name = LOCATIONS_MAP.get(lid, loc["label"])
    ws = wb_res.create_sheet()
    write_sheet(ws, reserved_headers, all_reserved[lid], name)

wb_res.save("bookings-reserved.xlsx")
print(f"\nbookings-reserved.xlsx saved ({sum(len(v) for v in all_reserved.values())} rows)")

# --- Write Blocked/Breaks Excel ---
blocked_headers = ["Event ID", "Profesional", "Tipo", "Titulo", "Descripcion", "Inicio", "Fin"]

wb_blk = Workbook()
wb_blk.remove(wb_blk.active)
for loc in locations:
    lid = loc["value"]
    name = LOCATIONS_MAP.get(lid, loc["label"])
    ws = wb_blk.create_sheet()
    write_sheet(ws, blocked_headers, all_blocked[lid], name)

wb_blk.save("bookings-blocked.xlsx")
print(f"bookings-blocked.xlsx saved ({sum(len(v) for v in all_blocked.values())} rows)")
